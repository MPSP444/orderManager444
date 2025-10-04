import os
import pandas as pd
from datetime import datetime
from PyQt5.QtWidgets import QMessageBox, QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel
from openpyxl import load_workbook
import pythoncom
import time
import traceback


class ExcelWriteErrorDialog(QDialog):
    """Диалоговое окно для обработки ошибок записи в Excel"""

    def __init__(self, parent=None, error_message=""):
        super().__init__(parent)
        self.retry_result = False
        self.initUI(error_message)

    def initUI(self, error_message):
        """Инициализация пользовательского интерфейса"""
        self.setWindowTitle('Ошибка записи в Excel')
        self.setGeometry(300, 300, 500, 200)

        layout = QVBoxLayout()

        # Отображение сообщения об ошибке
        error_label = QLabel(f"Не удалось записать данные в Excel файл.\n"
                             f"Возможно, файл открыт в другой программе.\n\n"
                             f"Ошибка: {error_message}\n\n"
                             f"Закройте файл и повторите операцию или пропустите запись.")
        error_label.setWordWrap(True)
        layout.addWidget(error_label)

        # Кнопки
        buttons_layout = QHBoxLayout()

        self.retry_btn = QPushButton('Повторить')
        self.retry_btn.clicked.connect(self.retry)

        self.skip_btn = QPushButton('Пропустить')
        self.skip_btn.clicked.connect(self.skip)

        buttons_layout.addStretch()
        buttons_layout.addWidget(self.retry_btn)
        buttons_layout.addWidget(self.skip_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

    def retry(self):
        """Обработчик нажатия кнопки 'Повторить'"""
        self.retry_result = True
        self.accept()

    def skip(self):
        """Обработчик нажатия кнопки 'Пропустить'"""
        self.retry_result = False
        self.accept()


class ExcelManager:
    """Класс для работы с Excel-файлами"""

    def __init__(self, parent=None):
        self.parent = parent
        from ui.windows.path_manager import PathManager
        self.path_manager = PathManager()

    def is_file_open(self, file_path):
        """Проверка, открыт ли файл другим процессом"""
        if not os.path.exists(file_path):
            return False

        try:
            # Пробуем открыть файл для записи
            with open(file_path, 'ab') as f:
                pass
            return False
        except PermissionError:
            return True
        except Exception:
            return True

    def add_payment_to_excel(self, payment_data):
        """
        Добавление данных об оплате в Excel файл

        payment_data - словарь с данными об оплате:
        {
            'order_id': ID заказа,
            'fio': ФИО клиента,
            'amount': сумма оплаты
        }
        """
        # Получение пути к Excel файлу
        excel_path = self.path_manager.get_path('payment_excel')
        print(f"Используемый путь к Excel: {excel_path}")

        if not excel_path:
            msg = "Путь к файлу Excel не указан. Пожалуйста, укажите путь в настройках."
            print(msg)
            QMessageBox.warning(self.parent, 'Ошибка', msg)
            return False

        # Фиксированные заголовки
        fixed_headers = ["ФИО", "Должность", "Наименование", "Дата", "Сумма", "Комментарий"]

        # Готовим данные
        current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        comment = f"Оплата по заказу #{payment_data['order_id']} от {payment_data['fio']}"

        # Фиксированные данные для записи
        row_data = [
            "Гурбанмурадов Мукам",  # ФИО всегда одинаковое
            "Программист",  # Должность всегда одинаковая
            "Заказы",  # Наименование всегда одинаковое
            current_datetime,  # Текущая дата и время
            payment_data['amount'],  # Сумма оплаты
            comment  # Комментарий
        ]

        # Попытка добавления данных в файл
        max_attempts = 3
        attempt = 0

        while attempt < max_attempts:
            try:
                # Проверяем, открыт ли файл
                is_open = self.is_file_open(excel_path)
                print(f"Файл открыт? {is_open}")

                if is_open:
                    dialog = ExcelWriteErrorDialog(
                        self.parent,
                        f"Файл '{excel_path}' открыт в другой программе.\nЗакройте файл и нажмите 'Повторить'."
                    )
                    dialog.exec_()

                    if dialog.retry_result:
                        attempt += 1
                        continue
                    else:
                        print("Пользователь выбрал пропустить запись")
                        return False

                # Загружаем существующий файл, если он есть
                if os.path.exists(excel_path):
                    try:
                        print("Открываем существующий файл...")
                        workbook = load_workbook(excel_path)
                        print(f"Доступные листы: {workbook.sheetnames}")

                        # Используем лист "Доходы" или первый лист
                        sheet_name = "Доходы"
                        if sheet_name not in workbook.sheetnames:
                            if len(workbook.sheetnames) > 0:
                                sheet_name = workbook.sheetnames[0]
                                print(f"Лист 'Доходы' не найден, используем '{sheet_name}'")
                            else:
                                sheet_name = "Доходы"
                                print("Не найдено ни одного листа, создаем 'Доходы'")
                                workbook.create_sheet(sheet_name)

                        sheet = workbook[sheet_name]

                        # Проверяем, есть ли заголовки
                        if sheet.max_row == 0:
                            # Добавляем заголовки если лист пустой
                            sheet.append(fixed_headers)

                        # Добавляем данные
                        print(f"Добавляем данные: {row_data}")
                        sheet.append(row_data)

                        # Сохраняем файл
                        print(f"Сохраняем файл: {excel_path}")
                        workbook.save(excel_path)
                        workbook.close()
                        print("Файл успешно сохранен")
                        return True

                    except PermissionError as e:
                        print(f"Ошибка доступа: {str(e)}")
                        dialog = ExcelWriteErrorDialog(
                            self.parent,
                            f"Доступ запрещен: {str(e)}\nВозможно, файл открыт в другой программе."
                        )
                        dialog.exec_()

                        if dialog.retry_result:
                            attempt += 1
                            time.sleep(1)
                            continue
                        else:
                            print("Пользователь выбрал пропустить запись")
                            return False

                    except Exception as e:
                        print(f"Ошибка при работе с файлом: {str(e)}")
                        traceback.print_exc()
                        dialog = ExcelWriteErrorDialog(self.parent, str(e))
                        dialog.exec_()

                        if dialog.retry_result:
                            attempt += 1
                            time.sleep(1)
                            continue
                        else:
                            print("Пользователь выбрал пропустить запись")
                            return False

                else:
                    # Если файла нет, создаем новый
                    try:
                        print("Создаем новый файл Excel...")
                        workbook = load_workbook()

                        # Удаляем стандартный лист (если есть)
                        if "Sheet" in workbook.sheetnames:
                            workbook.remove(workbook["Sheet"])

                        # Создаем лист "Доходы"
                        sheet = workbook.create_sheet("Доходы")

                        # Добавляем заголовки
                        sheet.append(fixed_headers)

                        # Добавляем данные
                        print(f"Добавляем данные: {row_data}")
                        sheet.append(row_data)

                        # Сохраняем файл
                        print(f"Сохраняем новый файл: {excel_path}")
                        workbook.save(excel_path)
                        workbook.close()
                        print("Новый файл успешно создан")
                        return True

                    except Exception as e:
                        print(f"Ошибка при создании нового файла: {str(e)}")
                        traceback.print_exc()
                        dialog = ExcelWriteErrorDialog(self.parent, str(e))
                        dialog.exec_()

                        if dialog.retry_result:
                            attempt += 1
                            time.sleep(1)
                            continue
                        else:
                            print("Пользователь выбрал пропустить запись")
                            return False

            except Exception as e:
                print(f"Общая ошибка, попытка {attempt + 1}: {str(e)}")
                traceback.print_exc()

                dialog = ExcelWriteErrorDialog(self.parent, str(e))
                dialog.exec_()

                if dialog.retry_result:
                    attempt += 1
                    time.sleep(1)
                    continue
                else:
                    print("Пользователь выбрал пропустить запись")
                    return False

        # Если все попытки неудачны
        msg = f'Не удалось записать данные в Excel файл после {max_attempts} попыток.'
        print(msg)
        QMessageBox.critical(self.parent, 'Ошибка', msg)
        return False
    def test_excel_connection(self, excel_path):
        """Тест подключения к Excel файлу"""
        try:
            if not os.path.exists(excel_path):
                return False, "Файл не существует"

            if self.is_file_open(excel_path):
                return False, "Файл открыт в другой программе"

            # Пробуем открыть файл через openpyxl
            workbook = load_workbook(excel_path)
            sheets = workbook.sheetnames
            workbook.close()

            return True, f"Соединение успешно. Листы в файле: {', '.join(sheets)}"

        except Exception as e:
            return False, f"Ошибка: {str(e)}"