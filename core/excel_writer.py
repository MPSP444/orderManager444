import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from ui.windows.path_manager import PathManager


class ExcelWriter:
    """Класс для работы с Excel-файлом доходов"""

    def __init__(self):
        self.path_manager = PathManager()

    def add_payment_record(self, fio, position, service_name, amount, comment=""):
        """
        Добавление записи об оплате в Excel-файл

        Args:
            fio (str): ФИО сотрудника
            position (str): Должность
            service_name (str): Наименование услуги
            amount (float): Сумма оплаты
            comment (str, optional): Комментарий. По умолчанию пустая строка.

        Returns:
            bool: Результат операции (True - успешно, False - ошибка)
        """
        try:
            # Получаем путь к Excel-файлу
            excel_path = self.path_manager.get_path('payment_excel')
            if not excel_path or not os.path.exists(excel_path):
                print(f"Ошибка: файл Excel не найден по пути {excel_path}")
                return False

            # Текущая дата
            current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Новая запись
            new_record = {
                'ФИО': fio,
                'Должность': position,
                'Наименование': service_name,
                'Дата': current_date,
                'Сумма': amount,
                'Комментарий': comment
            }

            # Попробуем загрузить файл с помощью pandas
            try:
                # Проверяем, существует ли файл
                if not os.path.exists(excel_path):
                    # Создаем новый DataFrame с нужными колонками
                    df = pd.DataFrame(columns=[
                        'ФИО', 'Должность', 'Наименование', 'Дата', 'Сумма', 'Комментарий'
                    ])
                else:
                    # Загружаем существующий файл
                    df = pd.read_excel(excel_path, sheet_name='Доходы')

                # Добавляем новую запись
                df = pd.concat([df, pd.DataFrame([new_record])], ignore_index=True)

                # Сохраняем обратно в файл
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Доходы', index=False)

                return True

            except Exception as e:
                print(f"Ошибка при работе с pandas: {str(e)}")

                # Альтернативный подход с использованием openpyxl
                try:
                    # Если файл не существует, создаем новый
                    if not os.path.exists(excel_path):
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = 'Доходы'

                        # Добавляем заголовки
                        ws.append([
                            'ФИО', 'Должность', 'Наименование', 'Дата', 'Сумма', 'Комментарий'
                        ])
                    else:
                        # Открываем существующий файл
                        wb = load_workbook(excel_path)

                        # Проверяем наличие листа 'Доходы'
                        if 'Доходы' not in wb.sheetnames:
                            ws = wb.create_sheet('Доходы')
                            ws.append([
                                'ФИО', 'Должность', 'Наименование', 'Дата', 'Сумма', 'Комментарий'
                            ])
                        else:
                            ws = wb['Доходы']

                    # Добавляем новую запись
                    ws.append([
                        fio, position, service_name, current_date, amount, comment
                    ])

                    # Сохраняем файл
                    wb.save(excel_path)
                    return True

                except Exception as e2:
                    print(f"Ошибка при работе с openpyxl: {str(e2)}")
                    return False

        except Exception as e:
            print(f"Общая ошибка при добавлении записи в Excel: {str(e)}")
            return False