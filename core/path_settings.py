import os
import json
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl import load_workbook, Workbook


class ExcelWriter:
    """Класс для работы с Excel-файлом доходов"""

    def __init__(self):
        self.settings_file = "paths_settings.json"

    def get_path(self, key):
        """Получение пути по ключу из настроек"""
        if os.path.exists(self.settings_file):
            try:
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    paths = json.load(f)
                    return paths.get(key, "")
            except Exception as e:
                print(f"Ошибка при загрузке путей: {str(e)}")
                return ""
        return ""

    def add_payment_record(self, fio, position, service_name, amount, comment=""):
        """
        Добавление записи об оплате в Excel-файл

        Args:
            fio (str): ФИО сотрудника
            position (str): Должность
            service_name (str): Наименование услуги
            amount (float): Сумма оплаты
            comment (str, optional): Комментарий. По умолчанию пустая строка.

        Returns:
            bool: Результат операции (True - успешно, False - ошибка)
        """
        try:
            # Получаем путь к Excel-файлу
            excel_path = self.get_path('payment_excel')
            if not excel_path:
                print("Ошибка: путь к Excel-файлу не настроен")
                return False

            # Проверяем существование папки назначения
            dest_dir = os.path.dirname(excel_path)
            if not os.path.exists(dest_dir):
                os.makedirs(dest_dir, exist_ok=True)

            # Текущая дата
            current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Новая запись
            new_record = {
                'ФИО': fio,
                'Должность': position,
                'Наименование': service_name,
                'Дата': current_date,
                'Сумма': amount,
                'Комментарий': comment
            }

            # Попробуем использовать pandas
            try:
                # Проверяем, существует ли файл
                if not os.path.exists(excel_path):
                    # Создаем новый DataFrame с нужными колонками
                    df = pd.DataFrame(columns=[
                        'ФИО', 'Должность', 'Наименование', 'Дата', 'Сумма', 'Комментарий'
                    ])
                else:
                    try:
                        # Загружаем существующий файл
                        df = pd.read_excel(excel_path, sheet_name='Доходы')
                    except Exception as sheet_error:
                        print(f"Ошибка при чтении листа 'Доходы': {sheet_error}")
                        # Если лист не найден, создаем новый DataFrame
                        df = pd.DataFrame(columns=[
                            'ФИО', 'Должность', 'Наименование', 'Дата', 'Сумма', 'Комментарий'
                        ])

                # Добавляем новую запись
                df = pd.concat([df, pd.DataFrame([new_record])], ignore_index=True)

                # Сохраняем обратно в файл
                if os.path.exists(excel_path):
                    # Если файл уже существует, сохраняем с сохранением других листов
                    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, sheet_name='Доходы', index=False)
                else:
                    # Если файл не существует, создаем новый
                    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='Доходы', index=False)

                print(f"Запись успешно добавлена в Excel: {excel_path}")
                return True

            except Exception as pandas_error:
                print(f"Ошибка при работе с pandas: {str(pandas_error)}")

                # Альтернативный подход с использованием только openpyxl
                try:
                    # Если файл не существует, создаем новый
                    if not os.path.exists(excel_path):
                        wb = Workbook()
                        ws = wb.active
                        ws.title = 'Доходы'

                        # Добавляем заголовки
                        ws.append([
                            'ФИО', 'Должность', 'Наименование', 'Дата', 'Сумма', 'Комментарий'
                        ])
                    else:
                        # Открываем существующий файл
                        wb = load_workbook(excel_path)

                        # Проверяем наличие листа 'Доходы'
                        if 'Доходы' not in wb.sheetnames:
                            ws = wb.create_sheet('Доходы')
                            ws.append([
                                'ФИО', 'Должность', 'Наименование', 'Дата', 'Сумма', 'Комментарий'
                            ])
                        else:
                            ws = wb['Доходы']

                    # Добавляем новую запись
                    ws.append([
                        fio, position, service_name, current_date, amount, comment
                    ])

                    # Сохраняем файл
                    wb.save(excel_path)
                    print(f"Запись успешно добавлена в Excel (openpyxl): {excel_path}")
                    return True

                except Exception as openpyxl_error:
                    print(f"Ошибка при работе с openpyxl: {str(openpyxl_error)}")
                    return False

        except Exception as e:
            print(f"Общая ошибка при добавлении записи в Excel: {str(e)}")
            return False